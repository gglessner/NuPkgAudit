#!/usr/bin/env python3
"""
XLSX Secrets Detection Module for UIPath Security Audit Tool
Scans all .xlsx files in packages for dangerous-sounding keys and their values

This module identifies potentially sensitive keys in Excel files and reports their values,
helping to find forgotten credentials, API keys, and other sensitive data in configuration files.
"""

import re
import sys
import logging
from pathlib import Path
from typing import List, Dict, Any, Tuple
import openpyxl
from openpyxl import load_workbook

# Add libraries path to sys.path to import config_helper
sys.path.insert(0, str(Path(__file__).parent.parent / 'libraries'))

try:
    from highlight_helper import highlight_match
except ImportError as e:
    print(f"Error importing libraries: {e}")
    sys.exit(1)

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Module description for the main audit tool
MODULE_DESCRIPTION = "Scans all .xlsx files in packages for dangerous-sounding keys containing credentials, passwords, API keys, and other sensitive data. Reports both the key names and their values."

# Dangerous key patterns - these indicate potentially sensitive data
DANGEROUS_KEY_PATTERNS = {
    # Password-related patterns
    'password': {
        'patterns': [
            r'.*pass.*word.*', r'.*password.*', r'.*pwd.*', r'.*passwd.*',
            r'.*pass_.*', r'.*_pass.*', r'.*pass$', r'^pass_.*',
            r'.*user.*pass.*', r'.*admin.*pass.*', r'.*login.*pass.*'
        ],
        'severity': 'HIGH',
        'description': 'Password-related key detected'
    },
    
    # Username/User ID patterns
    'username': {
        'patterns': [
            r'.*user.*name.*', r'.*username.*', r'.*user.*id.*', r'.*userid.*',
            r'.*user_.*', r'.*_user.*', r'.*login.*user.*', r'.*admin.*user.*',
            r'.*account.*', r'.*login.*', r'.*signin.*', r'.*logon.*'
        ],
        'severity': 'MEDIUM',
        'description': 'Username/User ID key detected'
    },
    
    # API Key patterns
    'api_key': {
        'patterns': [
            r'.*api.*key.*', r'.*apikey.*', r'.*api_key.*', r'.*key.*api.*',
            r'.*access.*key.*', r'.*secret.*key.*', r'.*private.*key.*',
            r'.*public.*key.*', r'.*auth.*key.*', r'.*client.*key.*'
        ],
        'severity': 'HIGH',
        'description': 'API key detected'
    },
    
    # Token patterns
    'token': {
        'patterns': [
            r'.*token.*', r'.*bearer.*', r'.*jwt.*', r'.*oauth.*',
            r'.*access.*token.*', r'.*refresh.*token.*', r'.*auth.*token.*',
            r'.*session.*token.*', r'.*csrf.*token.*'
        ],
        'severity': 'HIGH',
        'description': 'Authentication token detected'
    },
    
    # Secret patterns
    'secret': {
        'patterns': [
            r'.*secret.*', r'.*client.*secret.*', r'.*app.*secret.*',
            r'.*shared.*secret.*', r'.*private.*', r'.*confidential.*',
            r'.*sensitive.*', r'.*classified.*'
        ],
        'severity': 'HIGH',
        'description': 'Secret/private data detected'
    },
    
    # Connection string patterns
    'connection': {
        'patterns': [
            r'.*connection.*string.*', r'.*conn.*str.*', r'.*database.*conn.*',
            r'.*db.*conn.*', r'.*sql.*conn.*', r'.*server.*conn.*',
            r'.*connection.*', r'.*conn_.*', r'.*_conn.*'
        ],
        'severity': 'HIGH',
        'description': 'Database connection string detected'
    },
    
    # Certificate patterns
    'certificate': {
        'patterns': [
            r'.*cert.*', r'.*certificate.*', r'.*ssl.*cert.*', r'.*tls.*cert.*',
            r'.*x509.*', r'.*pem.*', r'.*p12.*', r'.*pfx.*', r'.*keystore.*'
        ],
        'severity': 'MEDIUM',
        'description': 'Certificate/SSL key detected'
    },
    
    # Encryption patterns
    'encryption': {
        'patterns': [
            r'.*encrypt.*key.*', r'.*decrypt.*key.*', r'.*cipher.*key.*',
            r'.*crypto.*key.*', r'.*hash.*key.*', r'.*salt.*', r'.*iv.*',
            r'.*initialization.*vector.*'
        ],
        'severity': 'HIGH',
        'description': 'Encryption key/parameter detected'
    },
    
    # Database patterns
    'database': {
        'patterns': [
            r'.*db.*pass.*', r'.*database.*pass.*', r'.*sql.*pass.*',
            r'.*db.*user.*', r'.*database.*user.*', r'.*sql.*user.*',
            r'.*sa.*pass.*', r'.*admin.*db.*', r'.*root.*pass.*'
        ],
        'severity': 'HIGH',
        'description': 'Database credential detected'
    },
    
    # Service account patterns
    'service_account': {
        'patterns': [
            r'.*service.*account.*', r'.*svc.*account.*', r'.*service.*user.*',
            r'.*svc.*user.*', r'.*service.*pass.*', r'.*svc.*pass.*',
            r'.*system.*account.*', r'.*app.*account.*'
        ],
        'severity': 'HIGH',
        'description': 'Service account credential detected'
    },
    
    # Cloud/Azure patterns
    'cloud': {
        'patterns': [
            r'.*azure.*key.*', r'.*aws.*key.*', r'.*gcp.*key.*', r'.*cloud.*key.*',
            r'.*subscription.*key.*', r'.*tenant.*id.*', r'.*client.*id.*',
            r'.*application.*id.*', r'.*resource.*id.*'
        ],
        'severity': 'HIGH',
        'description': 'Cloud service credential detected'
    },
    
    # Email/SMTP patterns
    'email': {
        'patterns': [
            r'.*smtp.*pass.*', r'.*email.*pass.*', r'.*mail.*pass.*',
            r'.*smtp.*user.*', r'.*email.*user.*', r'.*mail.*user.*',
            r'.*smtp.*auth.*', r'.*email.*auth.*'
        ],
        'severity': 'MEDIUM',
        'description': 'Email/SMTP credential detected'
    },
    
    # FTP/SFTP patterns
    'ftp': {
        'patterns': [
            r'.*ftp.*pass.*', r'.*ftp.*user.*', r'.*sftp.*pass.*', r'.*sftp.*user.*',
            r'.*ftp.*auth.*', r'.*sftp.*auth.*', r'.*ftp.*cred.*'
        ],
        'severity': 'MEDIUM',
        'description': 'FTP/SFTP credential detected'
    }
}

# Values that should be considered safe (not flagged as dangerous)
SAFE_VALUE_PATTERNS = [
    r'^$',  # Empty string
    r'^\s*$',  # Whitespace only
    r'^{x:Null}$',  # UiPath null
    r'^Nothing$',  # VB.NET null
    r'^null$',  # General null
    r'^<null>$',  # XML null
    r'^\[.*\]$',  # Variable references like [username]
    r'^.*\{.*\}.*$',  # String interpolation
    r'^.*Config.*$',  # Config references
    r'^.*config.*$',  # Config references (case insensitive)
    r'^.*in_.*$',  # Input variables
    r'^.*out_.*$',  # Output variables
    r'^.*io_.*$',  # Input/output variables
    r'^example.*$',  # Example values
    r'^sample.*$',  # Sample values
    r'^test.*$',  # Test values
    r'^demo.*$',  # Demo values
    r'^placeholder.*$',  # Placeholder values
    r'^template.*$',  # Template values
    r'^default.*$',  # Default values
    r'^dummy.*$',  # Dummy values
    r'^fake.*$',  # Fake values
    r'^mock.*$',  # Mock values
    r'^\*+$',  # Asterisks (password masks)
    r'^x+$',  # X characters (password masks)
    r'^-+$',  # Dashes
    r'^\.+$',  # Dots
    r'^true$',  # Boolean true
    r'^false$',  # Boolean false
    r'^0$',  # Zero
    r'^1$',  # One
]

def is_dangerous_key(key_name: str) -> Tuple[bool, str, str, str]:
    """
    Check if a key name matches dangerous patterns.
    
    Args:
        key_name: The key name to check
        
    Returns:
        Tuple of (is_dangerous, category, severity, description)
    """
    key_lower = key_name.lower().strip()
    
    for category, pattern_info in DANGEROUS_KEY_PATTERNS.items():
        patterns = pattern_info['patterns']
        severity = pattern_info['severity']
        description = pattern_info['description']
        
        for pattern in patterns:
            if re.search(pattern, key_lower, re.IGNORECASE):
                return True, category, severity, description
    
    return False, '', '', ''

def is_safe_value(value: str) -> bool:
    """
    Check if a value should be considered safe (not flagged).
    
    Args:
        value: The value to check
        
    Returns:
        bool: True if the value appears to be safe
    """
    if not value:
        return True
    
    value_str = str(value).strip()
    
    for pattern in SAFE_VALUE_PATTERNS:
        if re.search(pattern, value_str, re.IGNORECASE):
            return True
    
    return False

def scan_xlsx_file(file_path: Path) -> List[Dict[str, Any]]:
    """
    Scan a single .xlsx file for dangerous keys and their values.
    
    Args:
        file_path: Path to the .xlsx file
        
    Returns:
        List of issues found in the file
    """
    issues = []
    
    try:
        # Load the workbook
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        
        # Scan all worksheets
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Get all rows with data
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if not row or len(row) < 2:
                    continue
                
                # Check each cell pair (assuming key-value pairs)
                for col_idx in range(0, len(row) - 1, 2):
                    key_cell = row[col_idx]
                    value_cell = row[col_idx + 1] if col_idx + 1 < len(row) else None
                    
                    if not key_cell:
                        continue
                    
                    key_name = str(key_cell).strip()
                    value = str(value_cell).strip() if value_cell is not None else ""
                    
                    # Skip if key is too short or obviously not a key
                    if len(key_name) < 2 or key_name.isdigit():
                        continue
                    
                    # Check if key is dangerous
                    is_dangerous, category, severity, description = is_dangerous_key(key_name)
                    
                    if is_dangerous:
                        # Check if value is safe (variable reference, null, etc.)
                        if is_safe_value(value):
                            # Lower severity for safe values
                            if severity == 'HIGH':
                                severity = 'MEDIUM'
                            elif severity == 'MEDIUM':
                                severity = 'LOW'
                            description = f'{description} (safe value - likely variable reference)'
                        else:
                            # Check if value looks like actual sensitive data
                            if len(value) > 0 and not value.lower() in ['true', 'false', '0', '1']:
                                # This could be actual sensitive data
                                description = f'{description} - ACTUAL VALUE FOUND'
                        
                        # Create highlighted content
                        highlighted_key = highlight_match(key_name, key_name)
                        highlighted_value = highlight_match(value, value) if value else "(empty)"
                        content_line = f"{highlighted_key} = {highlighted_value}"
                        
                        issues.append({
                            'type': 'xlsx_secrets_detection',
                            'severity': severity,
                            'description': description,
                            'file': str(file_path),
                            'line': row_idx,
                            'line_content': content_line,
                            'full_line': f"{key_name} = {value}",
                            'matched_text': content_line,
                            'module': 'xlsx_secrets_detection',
                            'category': category,
                            'key_name': key_name,
                            'key_value': value,
                            'sheet_name': sheet_name,
                            'row': row_idx,
                            'column': col_idx + 1
                        })
            
            # Also check for key-value pairs in adjacent rows (vertical layout)
            for col_idx in range(worksheet.max_column):
                col_data = []
                for row in worksheet.iter_rows(min_col=col_idx+1, max_col=col_idx+1, values_only=True):
                    if row[0] is not None:
                        col_data.append((worksheet.cell(row=len(col_data)+1, column=col_idx+1).row, str(row[0]).strip()))
                
                # Look for key-value pairs in adjacent rows
                for i in range(0, len(col_data) - 1, 2):
                    key_row, key_name = col_data[i]
                    value_row, value = col_data[i + 1] if i + 1 < len(col_data) else (0, "")
                    
                    if len(key_name) < 2 or key_name.isdigit():
                        continue
                    
                    # Check if key is dangerous
                    is_dangerous, category, severity, description = is_dangerous_key(key_name)
                    
                    if is_dangerous:
                        # Check if value is safe
                        if is_safe_value(value):
                            if severity == 'HIGH':
                                severity = 'MEDIUM'
                            elif severity == 'MEDIUM':
                                severity = 'LOW'
                            description = f'{description} (safe value - likely variable reference)'
                        else:
                            if len(value) > 0 and not value.lower() in ['true', 'false', '0', '1']:
                                description = f'{description} - ACTUAL VALUE FOUND'
                        
                        # Avoid duplicates
                        duplicate = any(
                            issue['key_name'] == key_name and 
                            issue['key_value'] == value and 
                            issue['sheet_name'] == sheet_name
                            for issue in issues
                        )
                        
                        if not duplicate:
                            highlighted_key = highlight_match(key_name, key_name)
                            highlighted_value = highlight_match(value, value) if value else "(empty)"
                            content_line = f"{highlighted_key} = {highlighted_value}"
                            
                            issues.append({
                                'type': 'xlsx_secrets_detection',
                                'severity': severity,
                                'description': description,
                                'file': str(file_path),
                                'line': key_row,
                                'line_content': content_line,
                                'full_line': f"{key_name} = {value}",
                                'matched_text': content_line,
                                'module': 'xlsx_secrets_detection',
                                'category': category,
                                'key_name': key_name,
                                'key_value': value,
                                'sheet_name': sheet_name,
                                'row': key_row,
                                'column': col_idx + 1
                            })
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"Error scanning XLSX file {file_path}: {str(e)}")
        issues.append({
            'type': 'xlsx_secrets_detection',
            'severity': 'ERROR',
            'description': f'Error scanning XLSX file: {str(e)}',
            'file': str(file_path),
            'line': 0,
            'line_content': '',
            'module': 'xlsx_secrets_detection'
        })
    
    return issues

def scan_package(package_path: str, root_package_name: str = None, scanned_files: set = None) -> List[Dict[str, Any]]:
    """
    Scan a UIPath package for dangerous keys in all .xlsx files.
    
    Args:
        package_path: Path to the package directory
        root_package_name: Name of the root package directory
        scanned_files: Set of files that have already been scanned
        
    Returns:
        List of issues found
    """
    issues = []
    package = Path(package_path)
    
    # Initialize scanned_files if not provided
    if scanned_files is None:
        scanned_files = set()
    
    try:
        # Scan all .xlsx files recursively in the package
        for file_path in package.rglob('*.xlsx'):
            if file_path.is_file():
                # Skip if file has already been scanned
                if str(file_path) in scanned_files:
                    continue
                
                # Add file to scanned set
                scanned_files.add(str(file_path))
                
                logger.info(f"Scanning XLSX file: {file_path}")
                file_issues = scan_xlsx_file(file_path)
                if file_issues:
                    # Add package name to each issue
                    for issue in file_issues:
                        issue['package_name'] = root_package_name if root_package_name else package.name
                    issues.extend(file_issues)
                    
    except Exception as e:
        logger.error(f"Error scanning package {package_path}: {str(e)}")
        issues.append({
            'type': 'xlsx_secrets_detection',
            'severity': 'ERROR',
            'description': f'Error scanning package: {str(e)}',
            'file': str(package_path),
            'line': 0,
            'line_content': '',
            'module': 'xlsx_secrets_detection'
        })
    
    return issues

def find_line_number(content: str, search_text: str) -> int:
    """
    Find the line number of a specific text in content.
    
    Args:
        content: The text content to search in
        search_text: The text to search for
        
    Returns:
        Line number (1-based) where the text is found, or 0 if not found
    """
    lines = content.split('\n')
    for i, line in enumerate(lines, 1):
        if search_text in line:
            return i
    return 0 