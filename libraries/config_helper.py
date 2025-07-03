#!/usr/bin/env python3
"""
Configuration Helper Library
Provides helper functions for parsing configuration patterns and retrieving values from Config.xlsx files.

Author: Garland Glessner <gglessner@gmail.com>
License: GNU General Public License v3.0
Copyright (C) 2024 Garland Glessner

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import re
import logging
from pathlib import Path
from typing import Optional, Tuple
import openpyxl

logger = logging.getLogger(__name__)

def parse_in_config_pattern(value: str) -> Optional[Tuple[str, str]]:
    """
    Parse an In_Config pattern to extract the configuration key.
    
    Args:
        value: The value to parse (e.g., '[In_Config(&quot;CertPass&quot;).ToString]')
        
    Returns:
        Tuple of (config_key, original_value) if it's an In_Config pattern, None otherwise
    """
    # Pattern to match In_Config patterns with case variations
    # Matches: [In_Config("key").ToString], [in_config("key").ToString], etc.
    pattern = r'\[In_Config\s*\(\s*(?:&quot;|")([^"&]+)(?:&quot;|")\s*\)\s*\.\s*ToString\s*\]'
    
    match = re.search(pattern, value, re.IGNORECASE)
    if match:
        config_key = match.group(1)
        return config_key, value
    
    return None

def find_config_xlsx(package_path: Path) -> Optional[Path]:
    """
    Find Config.xlsx file by only checking package-directory/lib/net45/Data/Config.xlsx.
    If package_path is a file, start from its parent directory.
    Args:
        package_path: Path to the package directory or a subdirectory/file within it
    Returns:
        Path to Config.xlsx if found, None otherwise
    """
    p = package_path
    if p.is_file():
        p = p.parent
    p = p.resolve()

    # Only check default location: package-directory/lib/net45/Data/Config.xlsx
    default_config = p / 'lib' / 'net45' / 'Data' / 'Config.xlsx'
    if default_config.is_file():
        return default_config

    logger.warning(f"Config.xlsx not found in default location: {default_config}")
    return None

def get_config_value(config_file_path: Path, config_key: str) -> Optional[str]:
    """
    Retrieve a value from Config.xlsx file by key.
    
    Args:
        config_file_path: Path to the Config.xlsx file
        config_key: The configuration key to look up
        
    Returns:
        The configuration value if found, None otherwise
    """
    try:
        workbook = openpyxl.load_workbook(config_file_path, data_only=True)
        
        # Try to find the key in any worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Look for the key in the first column
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=2):
                if len(row) >= 2:
                    cell_key = row[0].value
                    cell_value = row[1].value
                    
                    if cell_key and str(cell_key).strip().lower() == config_key.lower():
                        return str(cell_value) if cell_value is not None else None
        
        logger.warning(f"Config key '{config_key}' not found in {config_file_path}")
        return None
        
    except Exception as e:
        logger.error(f"Error reading Config.xlsx file {config_file_path}: {str(e)}")
        return None

def extract_in_config_key_from_brackets(value: str) -> Optional[str]:
    """
    Extract the config key from any [ ... ] expression containing in_config("key") or in_config(&quot;key"), case-insensitive.
    Handles nesting, e.g. [DirectCast(in_config("My_pass"),SecureString)]
    Args:
        value: The value to parse (e.g., '[DirectCast(in_config("My_pass"),SecureString)]')
    Returns:
        The config key if found, else None
    """
    # Only process if value is in square brackets
    bracket_pattern = r'^\s*\[(.*)\]\s*$'
    m = re.match(bracket_pattern, value.strip())
    if not m:
        return None
    inner = m.group(1)
    # Look for in_config("key") or in_config(&quot;key") (case-insensitive)
    in_config_pattern = r'in_config\s*\(\s*(?:&quot;|")([^"&]+)(?:&quot;|")\s*\)'
    match = re.search(in_config_pattern, inner, re.IGNORECASE)
    if match:
        return match.group(1)
    return None

def resolve_in_config_value(value: str, package_path: Path) -> Optional[str]:
    """
    Resolve an In_Config or in_config pattern to its actual value from Config.xlsx.
    Handles both [In_Config("key").ToString] and [DirectCast(in_config("key"),...)] and similar.
    Args:
        value: The value to resolve (e.g., '[In_Config("CertPass").ToString]' or '[DirectCast(in_config("My_pass"),SecureString)]')
        package_path: Path to the package directory
    Returns:
        The resolved value if found, None otherwise
    """
    # Try original In_Config pattern
    parsed = parse_in_config_pattern(value)
    if parsed:
        config_key, _ = parsed
    else:
        # Try generic in_config inside brackets
        config_key = extract_in_config_key_from_brackets(value)
    if not config_key:
        return None
    # Find Config.xlsx file
    config_file = find_config_xlsx(package_path)
    if not config_file:
        logger.warning(f"Config.xlsx not found in package {package_path}")
        return None
    # Get the configuration value
    config_value = get_config_value(config_file, config_key)
    if config_value:
        logger.info(f"Resolved in_config '{config_key}' to value from {config_file}")
        return config_value
    return None 